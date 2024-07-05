import telebot
import requests
import os
from telebot import apihelper
from telebot import types
import yadisk
import time
import openpyxl
import random

# token from the bot
TOKEN = ''
# token from Yandex Disk
YANDEX_TOKEN = '' 

bot = telebot.TeleBot(TOKEN)
apihelper.API_URL = "http://localhost:8081/bot{0}/{1}"

y = yadisk.YaDisk(token=YANDEX_TOKEN)

local_path = 'D:/telegram-bot-api/'
bot_folder = 'bot_folder'
file_with_name_folder_event = 'name_folder_event.txt'
register_file = 'team_info.xlsx'
rate_file = 'rating.xlsx'
jury_file = "list_jury.xlsx"
topics_file = "topics.txt"

""" 
registration_data - two-dimensional array for storing registration data:
[chat_id]['id'] - team captain's ID
[chat_id]['team_name'] - team name
[chat_id]['full_name'] - full name of the captain
[chat_id]['user_link'] - username Telegram of the captain
[chat_id]['links_players'] - usernames Telegram of other participants
[chat_id]['group'] - group number
[chat_id]['remarks'] - text comment
"""
registration_data = {}
""" 
rate_data - a two-dimensional array for storing data on jury scores:
[chat_id]['shoot'] - score for shooting
[chat_id]['actor'] - score for acting skills
[chat_id]['editing'] - score for editing
[chat_id]['impression'] - score for overall impression
"""
rate_data = {}

""" 
get_name_folder_event - function to get the name of the current folder for the event
params: None
return: str
"""
def get_name_folder_event():
    file_path = local_path + file_with_name_folder_event
    disk_path = '/{}/{}'.format(bot_folder, file_with_name_folder_event)
    y.download(disk_path, file_path)
    file = open(file_path, 'r', encoding='utf-8')
    file_name = file.readline().strip('\n')
    return file_name

""" 
fill_rating_table - the function fills in a table with data about the jury's scores
As a result of the work, points for all criteria will be set in the table for each team
params: 
    team_names: array of strings 
    jury_names: array of strings
return: None
"""
def fill_rating_table(team_names, jury_names):
    folder = get_name_folder_event()
    file_path = local_path + rate_file
    disk_path = '/{}/{}'.format(folder, rate_file)
    y.download(disk_path, file_path)
    criteria = ['съемка', 'актерское мастерство', 'монтаж', 'общее впечатление']
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active    
   
    if ws.cell(2, 1).value != None:
        wb.close()
        return

    for i in range (len(jury_names)):
        ws.cell(row=1, column=i+3, value=jury_names[i])

    for i in range (len(team_names)):
        ws.cell(row=2+i*4, column=1, value=team_names[i])
        for j in range (len(criteria)):
          ws.cell(row=(2+i*4)+j, column=2, value=criteria[j])  
    
    wb.save(file_path)
    wb.close()
    y.remove(disk_path, permanently=True)
    y.upload(file_path, disk_path)

""" 
get_jury_info - function for getting two arrays with the full name of the jury and their Telegram usernames
params: None
return: 
    jury_names: array of strings 
    jury_usernames: array of strings
"""
def get_jury_info():
    folder = get_name_folder_event()
    file_path = local_path + jury_file
    disk_path = '/{}/{}'.format(folder, jury_file)
    y.download(disk_path, file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    jury_usernames = []
    jury_names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        jury_usernames.append(row[1])
        jury_names.append(row[0])
    wb.close()
    return jury_names, jury_usernames

""" 
get_theme_video - function to get the name of the theme for the video
params: None
return: str
"""
def get_theme_video():
    folder = get_name_folder_event()
    file_path = local_path + topics_file
    disk_path = '/{}/{}'.format(folder, topics_file)
    y.download(disk_path, file_path)
    with open(file_path, "r", encoding='utf-8') as topic_file:
        topic_list = [line.strip('\n') for line in topic_file.readlines()] 
        if len(topic_list) == 0:
            return None
        # жеребьевка темы
        num = random.randint(0, len(topic_list) - 1)
        theme = topic_list[num]
        topic_list.remove(topic_list[num])
    with open(file_path, "w", encoding='utf-8') as topic_file:   
        # перезаписываем файл с темами
        for line in topic_list:
            topic_file.write(line + '\n')
    y.remove(disk_path, permanently=True)
    y.upload(file_path, disk_path)
    return theme

""" 
get_user_link - function to get Telegram username 
params: 
    message: telebot.types.Message
return: str
"""
def get_user_link(message):
    user_id = message.chat.id
    user_link = f'@{message.from_user.username}'
    if None == user_link:
        user_link = 'tg://user?id=' + str(user_id) 
    return user_link

def send_user_link(message):
    user_id = message.chat.id
    user_link = f'<a href="tg://user?id={user_id}">tg://user?id={user_id}</a>'
    bot.send_message(message.chat.id, user_link, parse_mode='HTML')

@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.send_message(message.chat.id, "Привет! Для просмотра команд вызовите /menu")

""" 
get_info_about_teams - function returns two arrays with command names and an array with links to video commands
params: None
return: 
    team_names: array of strings 
    video_links: array of strings
"""
def get_info_about_teams():
    folder = get_name_folder_event()
    file_path = local_path + register_file
    disk_path = '/{}/{}'.format(folder, register_file)
    y.download(disk_path, file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    team_names = []
    video_links = []
    for i in range(2, ws.max_row + 1):
        name = ws.cell(i, 2).value
        link = ws.cell(i, 7).value
        if None != name and link != None:
            team_names.append(name)
            video_links.append(link)
    return team_names, video_links

""" 
handle_winner - the function determines the winner and sends a message with a list of winners to the chat
params: 
    message: telebot.types.Message
return: None
"""
def handle_winner(message):
    chat_id = message.chat.id
    folder = get_name_folder_event()
    file_path = local_path + rate_file
    disk_path = '/{}/{}'.format(folder, rate_file)
    y.download(disk_path, file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    points = []
    for i in range(2, ws.max_row - 2, 4):
        team_points = 0
        for j in range(3, ws.max_column + 1):
            p1 = ws.cell(i, j).value
            p2 = ws.cell(i + 1, j).value
            p3 = ws.cell(i + 2, j).value
            p4 = ws.cell(i + 3, j).value
           
            if p1 != None and p2 != None and p3 != None and p4 != None:
                team_points += int(p1)
                team_points += int(p2)
                team_points += int(p3)
                team_points += int(p4)
            else:
                bot.send_message(chat_id, "Не все оценки выставлены! Пожалуйста, убедитесь, что голосование окончено.")
                return
        points.append(team_points) 
    
    maxp = max(points)
    list = 'Победители: \n'
    for n in range(len(points)):
        if points[n] == maxp:
            list += ws.cell(2 + 4 * n, 1).value + '\n'
    
    bot.send_message(chat_id, list)

""" 
handle_rating - function to start the scoring process, sends a message to the request to rate according to the first criterion.
params: 
    message: telebot.types.Message
    team_num: int - current team number from the list
return: None
"""
def handle_rating(message, team_num):
    chat_id = message.chat.id
    global team_names
    global video_links
    if team_num == 0:
        team_names, video_links = get_info_about_teams()
        bot.send_message(chat_id, "Обрабатываю запрос...")
        # fill in the rating table if it is empty
        jury_names, jury_usernames = get_jury_info()
        fill_rating_table(team_names, jury_names)
    if team_num >= len(team_names):
        bot.send_message(chat_id, "Результаты сохранены", reply_markup=types.ReplyKeyboardRemove())
        return
    text = f'Команда: {team_names[team_num]}\n' 
    text += f'Ссылка на видео: {video_links[team_num]}'
    bot.send_message(chat_id, text)

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton("1")
    button2 = types.KeyboardButton("2")
    button3 = types.KeyboardButton("3")
    button4 = types.KeyboardButton("4")
    button5 = types.KeyboardButton("5")
    keyboard.row(button1, button2, button3, button4, button5)
    msg = bot.send_message(chat_id, "Оцените качество съемки:", reply_markup=keyboard)
    bot.register_next_step_handler(msg, rate_handler_shoot, team_num)

""" 
out_rating - function to fill in the table with the scores given by the jury
params: 
    message: telebot.types.Message
    team_num: int - current team number from the list
return: None
"""      
def out_rating(message, team_num):
    folder = get_name_folder_event()
    chat_id = message.chat.id
    username = '@' + message.chat.username
    file_path = local_path + rate_file
    disk_path = '/{}/{}'.format(folder, rate_file)
    y.download(disk_path, file_path)
    jury_names, jury_usernames = get_jury_info()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    row = 2 + 4 * team_num
    column = jury_usernames.index(username) + 3 # the jury is in the same order as in the table
    ws.cell(row=row, column=column, value=rate_data[chat_id]['shoot'])
    ws.cell(row=row + 1, column=column, value=rate_data[chat_id]['actor'])
    ws.cell(row=row + 2, column=column, value=rate_data[chat_id]['editing'])
    ws.cell(row=row + 3, column=column, value=rate_data[chat_id]['impression'])
    wb.save(file_path)
    wb.close()
    y.remove(disk_path, permanently=True)
    y.upload(file_path, disk_path)

    handle_rating(message, team_num + 1)

""" 
rate_handler_shoot - writes the score for the previous criterion to the array 
    and sends a message to the request for the next criterion.
params: 
    message: telebot.types.Message
    team_num: int - current team number from the list
return: None
"""  
def rate_handler_shoot(message, team_num):
    chat_id = message.chat.id
    rate_data[chat_id]={}
    rate_data[chat_id]["shoot"] = message.text
    msg = bot.send_message(chat_id, "Оцените актерское мастерство:")
    bot.register_next_step_handler(msg, rate_handler_actor, team_num)

def rate_handler_actor(message, team_num):
    chat_id = message.chat.id
    rate_data[chat_id]["actor"] = message.text
    msg = bot.send_message(chat_id, "Оцените качество монтажа:")
    bot.register_next_step_handler(msg, rate_handler_editing, team_num)

def rate_handler_editing(message, team_num):
    chat_id = message.chat.id
    rate_data[chat_id]["editing"] = message.text
    msg = bot.send_message(chat_id, "Оцените общее впечатление:")
    bot.register_next_step_handler(msg, rate_handler_impression, team_num)

def rate_handler_impression(message, team_num):
    rate_data[message.chat.id]["impression"] = message.text
    out_rating(message, team_num)
    
""" 
start_register - sends a message to the chat with the agreement to continue registration.
params: 
    message: telebot.types.Message
return: None
""" 
def start_register(message):
    chat_id = message.chat.id

    text = f"1. Сбор и обработка личных данных\n"
    text += f"   • Бот будет запрашивать у пользователей ФИО и контактные данные для идентификации и связи.\n"
    text += f"   • Пользователи должны ясно согласиться на обработку своих данных, включая использование видеороликов, предоставленных ими. \n"
    text += f"   • Согласие на обработку данных должно быть получено в форме явного согласия через кнопку в чате с ботом. \n"
    text += f"2. Права на использование видеороликов\n"
    text += f"   • Пользователи должны предоставить явное согласие на использование их видеороликов их хранение и пересылку.\n "
    text += f"   • Бот должен информировать пользователей о том, что их видеоролики могут быть использованы, и предоставить им возможность от этого отказаться. \n"
    text += f"3. Безопасность и конфиденциальность\n"
    text += f"   • Бот обеспечит безопасное хранение и обработку личных данных пользователей, используя современные методы шифрования. \n"
    text += f"   • Пользователи будут информированы о том, что их данные будут храниться и обрабатываться в соответствии с политикой конфиденциальности Telegram.\n"

    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton("Продолжить")
    button2 = types.KeyboardButton("Не продолжать")
    keyboard.add(button1, button2)
    bot.send_message(chat_id, "Продолжая, вы соглашаетесь на следующие правила пользования:\n" + text, reply_markup=keyboard)
    bot.register_next_step_handler(message, answer_handler)
    

def continue_register(message):
    chat_id = message.chat.id
    registration_data[chat_id] = {}
    registration_data[chat_id]['id'] = chat_id
    registration_data[chat_id]['user_link'] = get_user_link(message)
    msg = bot.send_message(chat_id, "Введите название команды:", reply_markup=types.ReplyKeyboardRemove())
    bot.register_next_step_handler(msg, get_team_name)


def answer_handler(message):
    if message.text =="Продолжить":
        continue_register(message)
    elif message.text =="Не продолжать":
        bot.send_message(message.chat.id, "К сожалению, вы не можете зарегистрироваться", reply_markup=types.ReplyKeyboardRemove())

def get_team_name(message):
    chat_id = message.chat.id
    registration_data[chat_id]['team_name'] = message.text
    msg = bot.send_message(chat_id, "Введите ФИО:")
    bot.register_next_step_handler(msg, get_links_players)

def get_links_players(message):    
    chat_id = message.chat.id   
    registration_data[chat_id]['full_name'] = message.text
    msg = bot.send_message(chat_id, "Введите ссылки на профили телеграм других участников команды в одном сообщении через пробел:")
    bot.register_next_step_handler(msg, get_group)

def get_group(message):    
    chat_id = message.chat.id   
    registration_data[chat_id]['links_players'] = message.text
    msg = bot.send_message(chat_id, "Введите номер группы:")
    bot.register_next_step_handler(msg, get_remarks)

def get_remarks(message):    
    chat_id = message.chat.id   
    registration_data[chat_id]['group'] = message.text
    msg = bot.send_message(chat_id, "Введите комментарий:")
    bot.register_next_step_handler(msg, end_register)

def end_register(message):    
    chat_id = message.chat.id   
    registration_data[chat_id]['remarks'] = message.text
    keyboard = types.InlineKeyboardMarkup()
    button1 = types.InlineKeyboardButton(text="Сохранить", callback_data='save_data')
    button2 = types.InlineKeyboardButton(text="Начать заново", callback_data='change_data')
    keyboard.add(button1, button2)
    bot.send_message(chat_id, "Вы хотите сохранить данные?", reply_markup=keyboard)

""" 
out_register_result - fills in the table with information about the team, 
    displays a message about the result and a message with a topic for the video in the chat.
params: 
    message: telebot.types.Message
return: None
"""
def out_register_result(message):
    chat_id = message.chat.id 
    folder = get_name_folder_event()
    file_path = local_path + register_file
    disk_path = '/{}/{}'.format(folder, register_file)

    y.download(disk_path, file_path)

    # fill in the table with information about the team
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active    
    row = 2
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == None:
            break
        row += 1
    ws.cell(row=row, column=1, value=registration_data[chat_id]['id'])
    ws.cell(row=row, column=2, value=registration_data[chat_id]['team_name'])
    ws.cell(row=row, column=3, value=registration_data[chat_id]['full_name'])
    ws.cell(row=row, column=4, value=registration_data[chat_id]['user_link'])
    ws.cell(row=row, column=5, value=registration_data[chat_id]['links_players'])
    ws.cell(row=row, column=6, value=registration_data[chat_id]['group'])
    ws.cell(row=row, column=8, value=registration_data[chat_id]['remarks'])
    wb.save(file_path)
    wb.close()
    
    y.remove(disk_path, permanently=True)
    y.upload(file_path, disk_path)
    
    bot.send_message(chat_id, 'Данные сохранены')
    team_info = f"Название команды: {registration_data[chat_id]['team_name']}\n"
    team_info += f"ФИО: {registration_data[chat_id]['full_name']}\n"
    team_info += f"Ссылки на других участников команды: {registration_data[chat_id]['links_players']}\n"
    team_info += f"Номер группы: {registration_data[chat_id]['group']}\n"
    team_info += f"Комментарий: {registration_data[chat_id]['remarks']}"
    bot.send_message(chat_id, team_info)

    bot.send_message(chat_id, "Выбираю тему для видео...") 
    # give a theme for the video
    theme = get_theme_video()
    if theme == None:
       bot.send_message(chat_id, "Упс, темы кончились...") 
    else:
        bot.send_message(chat_id, "Ваша тема: " + theme)

  
@bot.callback_query_handler(func=lambda call: call.data == 'save_data')
def save_btn(call):
    out_register_result(call.message)
    
@bot.callback_query_handler(func=lambda call: call.data == 'change_data')
def change_btn(call):
    start_register(call.message)

@bot.message_handler(content_types=['video'])
def handle_video(message):
    chat_id = message.chat.id 
    file_id = message.video.file_id
    bot.send_message(chat_id, "Загружаю файл...")
    # uploading the file to Yandex Disk
    while True:
        try:
            if bot.get_file(file_id).file_path:
                break
            else:
                time.sleep(10)
        except requests.exceptions.ReadTimeout:
            time.sleep(10)
            pass
  
    file_info = bot.get_file(file_id)
    file_path = file_info.file_path
    file_name = os.path.basename(file_info.file_path)
    folder = get_name_folder_event()
    try:
        y.upload(file_path, '/{}/{}'.format(folder, file_name))
    except yadisk.exceptions.PathExistsError:
        pass
    
    if y.exists(f'/{folder}/{file_name}'):
        bot.send_message(message.chat.id, "Видео успешно загружено")
    else:
        bot.send_message(message.chat.id, "Произошла ошибка при загрузке видео")
    # make it public and get the href
    time.sleep(10)
    resource = y.publish('/{}/{}'.format(folder, file_name))
    # get metadata to pick up the public link
    video_link = y.get_meta('/{}/{}'.format(folder, file_name)).FIELDS["public_url"]

    # adding a link to excel
    f_path = local_path + register_file
    disk_path = '/{}/{}'.format(folder, register_file)
    y.download(disk_path, f_path)

    wb = openpyxl.load_workbook(f_path)
    ws = wb.active

    i = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == chat_id:
            break 
        i += 1
    ws.cell(row=i, column=7, value=video_link)

    wb.save(f_path)
    wb.close()
    
    y.remove(disk_path, permanently=True)
    y.upload(f_path, disk_path)

@bot.message_handler(commands=['menu'])
def handle_menu(message):
    user_link = f'@{message.from_user.username}'
    keyboard = types.InlineKeyboardMarkup()
    jury_names, jury_usernames = get_jury_info()
    if user_link in jury_usernames:
        keyboard.add(types.InlineKeyboardButton("Поставить оценки", callback_data='rate'))
        keyboard.add(types.InlineKeyboardButton("Получить победителя", callback_data='get_winner'))
    else:
        keyboard.add(types.InlineKeyboardButton("Зарегистрировать команду", callback_data='register'))
        keyboard.add(types.InlineKeyboardButton("Загрузить видео", callback_data='video'))

    bot.send_message(message.chat.id, "Меню", reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: call)
def menu_btn(call):
    message = call.message 
    if call.data == 'register':
        start_register(message)
    elif call.data == 'video':
        bot.send_message(message.chat.id, "Отправьте видео в чат")
    elif call.data == 'rate':
        handle_rating(message, 0)
    elif call.data == 'get_winner':
        handle_winner(message)


bot.polling()